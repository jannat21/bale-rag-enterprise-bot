import openpyxl
import re
import csv

def extract_info(cell_value):
    """اطلاعات موسسه (نام، آدرس، تلفن) را از سلول A استخراج می‌کند."""
    if not cell_value:
        return None, None, None
    lines = str(cell_value).split('\n')
    name = lines[0].strip()
    address = ''
    phone = ''
    for line in lines[1:]:
        line = line.strip()
        if not line:
            continue
        if 'شماره تماس' in line or 'تلفن' in line or 'شماره' in line:
            phone_numbers = re.findall(r'[\d\-]+', line)
            phone = ' / '.join(phone_numbers) if phone_numbers else line
        else:
            address = line if not address else address + ' - ' + line
    return name, address, phone

def parse_price(cell_value):
    if not cell_value:
        return ''
    return str(cell_value).strip()

def is_title_cell(cell_value):
    if not cell_value:
        return False
    val = str(cell_value).strip()
    if 'تومان' in val:
        return False
    if re.search(r'[\d,\.]+', val):
        if re.match(r'^[\d,\.]+$', val.replace(' ', '')):
            return False
    return True

def get_grade_and_age(column_idx):
    grade_map = {
        2: ('پیش دبستانی', 4, 6),
        3: ('دبستان1', 6, 9),
        4: ('دبستان2', 9, 12),
        5: ('متوسطه1', 12, 15),
        6: ('متوسطه2', 15, 18),
        7: ('بزرگسالان', 19, 99),
    }
    return grade_map.get(column_idx, ('', 0, 0))

# بارگذاری فایل
wb = openpyxl.load_workbook('part2.xlsx')
sheet = wb['Sheet1']

output_rows = []
current_institute = None
current_address = None
current_phone = None
i = 1

while i <= sheet.max_row:
    cell_a = sheet.cell(row=i, column=1).value
    if cell_a and isinstance(cell_a, str) and any(k in cell_a for k in ['اتاق بازرگانی', 'جهاد دانشگاهی', 'مجتمع آموزشی', 'اداره کل', 'تاسیسات', 'مرکز تخصصی', 'آموزش رباتیک', 'شهرک مشاغل', 'موسسه تیزفکری', 'سازمان فرهنگی', 'مجموعه آموزشی ستاک', 'مرکز نجوم', 'شرکت مانا', 'مجموعه ورزشی خلیج فارس', 'مدرسه فوتبال', 'کانون ورزشی', 'مجموعه فرهنگی ورزشی', 'باشگاه پالاس', 'زبان صدوقی', 'باشگاه ورزشی شاهین', 'پردیس هنری', 'خانه مادر', 'مجموعه تخصصی نارنج', 'خانه ژیمناستیک', 'باشگاه ورزشی علمشیری', 'استخر ابوذر', 'استخر نصر', 'باشگاه فرهنگی ورزشی تام', 'استخر صبا', 'آکادمی تنیس', 'مجموعه ورزشی تختی', 'کلینیک تخصصی', 'باشگاه ورزشی اوتانا', 'آموزشگاه فنی و حرفه ای قله', 'استخر قصر موج', 'موسسه آموزشی تکین', 'استخر شاهین', 'باشگاه ورزشی شهدای مخابرات', 'خانه شطرنج', 'مرکز رشد فطرت', 'آموزش تنیس خاکی']):
        name, address, phone = extract_info(cell_a)
        current_institute = name
        current_address = address
        current_phone = phone
        i += 1
        continue

    if current_institute:
        for col in range(2, 8):
            title_cell = sheet.cell(row=i, column=col).value
            if is_title_cell(title_cell):
                title = str(title_cell).strip()
                price_cell = sheet.cell(row=i+1, column=col).value
                price = parse_price(price_cell)
                grade, min_age, max_age = get_grade_and_age(col)
                if title and price:
                    output_rows.append({
                        'institute': current_institute,
                        'address': current_address or '',
                        'phone': current_phone or '',
                        'title': title,
                        'price': price,
                        'min_age': min_age,
                        'max_age': max_age,
                        'grade': grade
                    })
        i += 1
    else:
        i += 1

# نوشتن فایل CSV با encoding utf-8-sig (سازگار با اکسل)
with open('output.csv', 'w', encoding='utf-8-sig', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['institute', 'address', 'phone', 'title', 'price', 'min_age', 'max_age', 'grade'])
    for row in output_rows:
        writer.writerow([row['institute'], row['address'], row['phone'], row['title'], row['price'], row['min_age'], row['max_age'], row['grade']])

print("فایل output.csv با موفقیت ایجاد شد.")