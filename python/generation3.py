import openpyxl
import csv
import re

def clean_filename(name):
    """فقط برای نام فایل استفاده می‌شود، اینجا نیازی نیست"""
    return name

def clean_price(text):
    """حذف نقطه و کاما از قیمت"""
    if not text:
        return "قیمت موجود نیست"
    cleaned = text.replace('.', '').replace(',', '')
    cleaned = ' '.join(cleaned.split())
    return cleaned

def extract_courses_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    # خواندن سرستون‌ها (مقطع و سن)
    col_levels = {}
    col_ages = {}
    for col_idx in range(2, 8):  # B تا G
        col_levels[col_idx] = str(sheet.cell(row=1, column=col_idx).value or "").strip()
        col_ages[col_idx] = str(sheet.cell(row=2, column=col_idx).value or "").strip()

    # خواندن تمام سطرها از سطر سوم به بعد
    all_rows = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        all_rows.append(row)

    centers = []
    i = 0
    while i < len(all_rows):
        row = all_rows[i]

        # شروع موسسه جدید (ستون A غیرخالی)
        if row[0] is not None and str(row[0]).strip() != "":
            center_name = str(row[0]).strip()
            i += 1

            # جمع‌آوری سلول‌های غیرخالی از ستون‌های B تا G برای این موسسه
            cells = []
            while i < len(all_rows):
                current_row = all_rows[i]

                # بررسی پایان موسسه
                # 1. اگر ستون A غیرخالی باشد → موسسه جدید
                if current_row[0] is not None and str(current_row[0]).strip() != "":
                    break
                # 2. اگر به بخش جدول یارانه رسیدیم
                if current_row[1] is not None and ("مبلغ" in str(current_row[1]) or "خانواده" in str(current_row[1])):
                    break

                # جمع‌آوری سلول‌های غیرخالی در ستون‌های B تا G
                for col_idx in range(2, 8):
                    val = current_row[col_idx - 1]
                    if val is not None and str(val).strip() != "":
                        cells.append({
                            'col': col_idx,
                            'value': str(val).strip()
                        })
                i += 1

            # پردازش هر ستون به‌صورت جداگانه برای جفت‌سازی عنوان و قیمت
            courses = []
            for col_idx in range(2, 8):
                # فیلتر سلول‌های مربوط به این ستون
                col_cells = [c for c in cells if c['col'] == col_idx]
                values = [c['value'] for c in col_cells]

                j = 0
                while j < len(values):
                    current = values[j]
                    next_val = values[j + 1] if j + 1 < len(values) else None

                    # حالت 1: عنوان + قیمت (معمول)
                    if "تومان" not in current and next_val is not None and "تومان" in next_val:
                        courses.append({
                            'title': current,
                            'price': clean_price(next_val),
                            'level': col_levels[col_idx],
                            'age': col_ages[col_idx]
                        })
                        j += 2

                    # حالت 2: قیمت + عنوان (برعکس)
                    elif "تومان" in current and next_val is not None and "تومان" not in next_val:
                        courses.append({
                            'title': next_val,
                            'price': clean_price(current),
                            'level': col_levels[col_idx],
                            'age': col_ages[col_idx]
                        })
                        j += 2

                    # حالت 3: عنوان بدون قیمت
                    elif "تومان" not in current:
                        courses.append({
                            'title': current,
                            'price': "قیمت موجود نیست",
                            'level': col_levels[col_idx],
                            'age': col_ages[col_idx]
                        })
                        j += 1

                    # حالت 4: قیمت بدون عنوان (نادیده گرفته می‌شود)
                    else:
                        j += 1

            centers.append({
                'name': center_name,
                'courses': courses
            })

        else:
            i += 1

    return centers


def save_to_csv(centers, output_file="courses_table.csv"):
    """ذخیره در فایل CSV با ستون‌های خواسته شده"""
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        # هدر
        writer.writerow(['نام موسسه', 'عنوان دوره', 'مبلغ دوره', 'بازه سنی', 'مقطع تحصیلی'])

        for center in centers:
            for course in center['courses']:
                writer.writerow([
                    center['name'],
                    course['title'],
                    course['price'],
                    course['age'],
                    course['level']
                ])


if __name__ == "__main__":
    excel_file = "part2.xlsx"
    centers = extract_courses_from_excel(excel_file)
    save_to_csv(centers, "courses_table.csv")

    print(f"✅ تعداد موسسات استخراج‌شده: {len(centers)}")
    total_courses = sum(len(c['courses']) for c in centers)
    print(f"✅ تعداد کل دوره‌ها: {total_courses}")
    print("✅ فایل courses_table.csv با موفقیت ایجاد شد.")